from django.conf import settings
from django.contrib import messages
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth.decorators import login_required
from django.core.paginator import EmptyPage, PageNotAnInteger, Paginator
from django.db.models import Q
from django.shortcuts import get_object_or_404, redirect, render

from ..models import (
    Attendance,
    Certificate,
    Cohort,
    Course,
    Enrollment,
    Registration,
    Student,
    StudentGrades,
    Submission,
)


def index(request):
    # Staff members go to cohorts page as their home
    if request.user.is_staff or request.user.is_superuser:
        return redirect("cohort_list")
    elif request.user.is_authenticated:
        return redirect("dashboard")

    return render(request, "app/public_home.html")


def privacy(request):
    return render(request, "app/privacy.html")

def terms(request):
    return render(request, "app/terms.html")


@login_required
def dashboard(request):
    """Main dashboard view - public home or authenticated dashboard"""
    # Regular users see unified dashboard
    from django.db.models import Prefetch

    # Get or create student profile for logged-in user
    try:
        student = Student.objects.prefetch_related(
            Prefetch(
                "registrations",
                queryset=Registration.objects.select_related("cohort")
                .prefetch_related(
                    Prefetch(
                        "enrollments",
                        queryset=Enrollment.objects.filter(
                            status__in=["IN_PROGRESS", "COMPLETED"]
                        )
                        .select_related("course", "registration", "certificate")
                        .prefetch_related(
                            "course__assignments", "submissions__assignment"
                        )
                        .order_by("course__name"),
                    )
                )
                .order_by("-cohort__start_date"),
            )
        ).get(user=request.user)
    except Student.DoesNotExist:
        # Try to find existing student by email (from Google Classroom sync or admin creation)
        try:
            student = Student.objects.prefetch_related(
                Prefetch(
                    "registrations",
                    queryset=Registration.objects.select_related("cohort")
                    .prefetch_related(
                        Prefetch(
                            "enrollments",
                            queryset=Enrollment.objects.filter(
                                status__in=["IN_PROGRESS", "COMPLETED"]
                            )
                            .select_related("course", "registration", "certificate")
                            .prefetch_related(
                                "course__assignments", "submissions__assignment"
                            )
                            .order_by("course__name"),
                        )
                    )
                    .order_by("-cohort__start_date"),
                )
            ).get(email=request.user.email)
            # Link this existing student to the user account
            student.user = request.user
            student.save()
        except Student.DoesNotExist:
            # No student profile - show available cohorts for registration
            # Student profile will be auto-created when they register for a cohort
            open_cohorts = Cohort.objects.filter(
                is_open_for_registration=True
            ).order_by("-start_date")

            # Build cohort data with registration status
            available_cohorts_data = []
            for cohort in open_cohorts:
                available_cohorts_data.append(
                    {
                        "cohort": cohort,
                        "can_register": cohort.can_accept_registrations,
                        "current_count": cohort.total_enrolled_students,
                    }
                )

            return render(
                request,
                "app/registration_home.html",
                {
                    "user": request.user,
                    "available_cohorts_data": available_cohorts_data,
                },
            )

    # Get approved and pending registrations
    approved_registrations = (
        Registration.objects.filter(student=student, status="APPROVED")
        .select_related("cohort")
        .order_by("-created_at")
    )

    pending_registrations = (
        Registration.objects.filter(student=student, status="PENDING")
        .select_related("cohort")
        .order_by("-created_at")
    )

    # Get enrolled course IDs
    enrolled_course_ids = Enrollment.objects.filter(
        registration__student=student, status__in=["IN_PROGRESS", "COMPLETED"]
    ).values_list("course_id", flat=True)

    # Get available courses (flat list, no cohort grouping)
    available_courses = []
    primary_cohort = None
    primary_cohort_enrollment_count = 0
    if approved_registrations.exists():
        # Use the first approved registration's cohort for enrollment
        primary_registration = approved_registrations.first()
        primary_cohort = primary_registration.cohort

        # Count enrollments in the primary cohort only
        primary_cohort_enrollment_count = Enrollment.objects.filter(
            registration=primary_registration, status__in=["IN_PROGRESS", "COMPLETED"]
        ).count()

        # Show ALL visible courses, not just ones with existing enrollments
        available_courses = (
            Course.objects.filter(is_visible=True, course_state="ACTIVE")
            .exclude(
                id__in=enrolled_course_ids  # Exclude already enrolled courses
            )
            .prefetch_related("assignments")
            .order_by("name")
        )

    # Get cohorts available for registration
    open_cohorts = Cohort.objects.filter(is_open_for_registration=True).order_by(
        "-start_date"
    )

    # Get student's existing registrations
    existing_registration_cohort_ids = Registration.objects.filter(
        student=student
    ).values_list("cohort_id", flat=True)

    # Build available cohorts data
    available_cohorts_data = []
    for cohort in open_cohorts:
        is_registered = cohort.id in existing_registration_cohort_ids
        registration = None
        if is_registered:
            registration = Registration.objects.filter(
                student=student, cohort=cohort
            ).first()

        available_cohorts_data.append(
            {
                "cohort": cohort,
                "is_registered": is_registered,
                "registration": registration,
                "can_register": cohort.can_accept_registrations and not is_registered,
                "current_count": cohort.total_enrolled_students,
            }
        )

    # Check if student can mark attendance this week
    from datetime import date

    today = date.today()

    # Check if already marked for today and get hours
    today_attendance = Attendance.objects.filter(student=student, date=today).first()

    marked_today = today_attendance is not None
    hours_today = today_attendance.hours_spent if today_attendance else 0

    context = {
        "student": student,
        "student_name": student.full_name,
        "student_email": student.email,
        "enrollments": Enrollment.objects.filter(
            registration__student=student,
            course__is_visible=True,
            status__in=["IN_PROGRESS", "COMPLETED"],
        ),
        "total_enrollments": student.enrollment_count,
        "primary_cohort_enrollment_count": primary_cohort_enrollment_count,
        "avg_completion": student.average_completion_rate,
        "avg_assignment": student.average_score,
        "avg_improvement": student.average_improvement,
        "has_improvement": student.has_improvement_data,
        "avg_on_time": student.average_on_time_rate,
        "available_courses": available_courses,
        "primary_cohort": primary_cohort,
        "has_pending_registration": pending_registrations.exists(),
        "available_cohorts_data": available_cohorts_data,
        "can_mark_attendance": student.has_active_registration,
        "marked_today": marked_today,
        "hours_today": hours_today,
        "attendance_rate": student.attendance_rate,
        "total_hours": student.total_attendance_hours,
    }
    return render(request, "app/dashboard.html", context)


@staff_member_required
def reports(request):
    """Reports view showing student enrollments - requires staff access"""
    import csv
    from django.http import HttpResponse
    from django.db.models import Prefetch
    from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger

    # Optimize query with prefetching to avoid N+1 queries
    # - Submissions are needed for enrollment.overall_average_score
    # - Attendance is needed for registration.session_attendance_rate
    # - Assignments are needed for calculating scores
    enrollments = (
        Enrollment.objects.select_related(
            "registration__student",
            "course",
            "registration",
            "registration__cohort",
            "certificate",
        )
        .prefetch_related(
            Prefetch(
                "submissions",
                queryset=Submission.objects.select_related("assignment").filter(
                    assigned_grade__isnull=False,
                    assignment__max_points__isnull=False,
                    assignment__max_points__gt=0,
                ),
            ),
            Prefetch(
                "registration__student__attendance",
                queryset=Attendance.objects.select_related("cohort"),
            ),
            # "course__assignments",
        )
        .filter(course__is_visible=True)
        # .order_by("registration__student__full_name", "cohort__name", "course__name")
    )

    # Calculate unique student count
    unique_student_count = enrollments.values("registration__student").distinct().count()

    # Handle Excel export
    if request.GET.get("format") == "excel":
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "Student Enrollments"

        # Write header row with styling
        headers = [
            "Student Name",
            "Email",
            "Cohort",
            "Course",
            "Attendance %",
            "Grade %",
            "Certificate",
        ]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color="4F46E5", end_color="4F46E5", fill_type="solid"
            )
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Track student groupings for merging
        row = 2
        current_student_id = None
        student_start_row = None

        for enrollment in enrollments:
            attendance = (
                f"{enrollment.registration.session_attendance_rate:.1f}"
                if enrollment.registration
                else "N/A"
            )
            grade = (
                f"{enrollment.overall_average_score:.1f}"
                if enrollment.overall_average_score is not None
                else "N/A"
            )
            certificate = "Issued" if enrollment.has_certificate else "Declined"

            # Check if this is a new student
            if enrollment.student.id != current_student_id:
                # Merge previous student cells if needed
                if current_student_id is not None and student_start_row < row - 1:
                    ws.merge_cells(
                        start_row=student_start_row,
                        start_column=1,
                        end_row=row - 1,
                        end_column=1,
                    )
                    ws.merge_cells(
                        start_row=student_start_row,
                        start_column=2,
                        end_row=row - 1,
                        end_column=2,
                    )
                    # Center the merged cells
                    ws.cell(row=student_start_row, column=1).alignment = Alignment(
                        vertical="center"
                    )
                    ws.cell(row=student_start_row, column=2).alignment = Alignment(
                        vertical="center"
                    )

                # Start new student group
                current_student_id = enrollment.student.id
                student_start_row = row

                # Write student name and email
                ws.cell(row=row, column=1, value=enrollment.student.full_name)
                ws.cell(row=row, column=2, value=enrollment.student.email)

            # Write course data
            ws.cell(row=row, column=3, value=enrollment.cohort.name)
            ws.cell(row=row, column=4, value=enrollment.course.name)
            ws.cell(row=row, column=5, value=attendance)
            ws.cell(row=row, column=6, value=grade)
            ws.cell(row=row, column=7, value=certificate)

            row += 1

        # Merge last student cells if needed
        if current_student_id is not None and student_start_row < row - 1:
            ws.merge_cells(
                start_row=student_start_row,
                start_column=1,
                end_row=row - 1,
                end_column=1,
            )
            ws.merge_cells(
                start_row=student_start_row,
                start_column=2,
                end_row=row - 1,
                end_column=2,
            )
            ws.cell(row=student_start_row, column=1).alignment = Alignment(
                vertical="center"
            )
            ws.cell(row=student_start_row, column=2).alignment = Alignment(
                vertical="center"
            )

        # Auto-size columns
        for col in range(1, 8):
            ws.column_dimensions[get_column_letter(col)].width = 20

        # Create response
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = (
            'attachment; filename="student_enrollments_report.xlsx"'
        )
        wb.save(response)

        return response

    # Paginate enrollments (20 per page)
    paginator = Paginator(enrollments, settings.PER_PAGE)
    page = request.GET.get("page", 1)

    try:
        enrollments_page = paginator.page(page)
    except PageNotAnInteger:
        enrollments_page = paginator.page(1)
    except EmptyPage:
        enrollments_page = paginator.page(paginator.num_pages)

    # Calculate rowspan for student cells (for merged cells in template)
    enrollments_list = list(enrollments_page)

    context = {
        "enrollments": enrollments_list,
        "total_enrollments": enrollments.count(),
        "total_students": unique_student_count,
        "page_obj": enrollments_page,
    }
    return render(request, "app/reports.html", context)


def student_grades(request):
    # Get search query
    search_query = request.GET.get("q", "").strip()

    current = Cohort.objects.filter(status='ACTIVE').first()
    grades = StudentGrades.objects.filter(cohort=current.name).all()

    if search_query:
        grades = grades.filter(
            Q(student__icontains=search_query) | Q(email__icontains=search_query)
        )

    # Paginate students (20 per page)
    paginator = Paginator(grades, settings.PER_PAGE)
    page = request.GET.get("page", 1)

    try:
        grades_page = paginator.page(page)
    except PageNotAnInteger:
        grades_page = paginator.page(1)
    except EmptyPage:
        grades_page = paginator.page(paginator.num_pages)

    return render(request, "app/student_grades.html", {
        "grades": grades_page,
        "cohort": current,
        "search_query": search_query,
    })

@login_required
def issues(request):
    """Issues landing page - show all issue categories"""

    # For now, no automatic issues since Attendance uses proper FKs
    # Future issue detection can be added here

    issue_categories = [
        # Future issue types can be added here
        # {
        #     'title': 'Duplicate Enrollments',
        #     'description': 'Students enrolled multiple times in the same course',
        #     'count': 0,
        #     'url': 'issues_duplicate_enrollments',
        #     'icon': 'users',
        #     'severity': 'warning',
        # },
    ]

    context = {
        "issue_categories": issue_categories,
        "total_issues": sum(cat["count"] for cat in issue_categories),
    }
    return render(request, "app/issues.html", context)


@login_required
def enroll_in_course(request, course_id):
    """Enroll student in a course within their cohort"""
    if request.method != "POST":
        return redirect("home")

    # Get cohort_id from POST data
    cohort_id = request.POST.get("cohort_id")
    if not cohort_id:
        messages.error(request, "Cohort information missing.")
        return redirect("home")

    # Get student
    try:
        student = Student.objects.get(user=request.user)
    except Student.DoesNotExist:
        messages.error(request, "Student profile not found.")
        return redirect("home")

    # Get course and cohort
    course = get_object_or_404(Course, id=course_id)
    cohort = get_object_or_404(Cohort, id=cohort_id)

    # Verify student has approved registration for this cohort
    registration = Registration.objects.filter(
        student=student, cohort=cohort, status="APPROVED"
    ).first()

    if not registration:
        messages.error(request, f"You are not registered for the {cohort.name} cohort.")
        return redirect("home")

    # Check if already enrolled
    existing = Enrollment.objects.filter(
        registration__student=student, course=course
    ).first()

    if existing:
        existing.status = Enrollment.StatusChoices.IN_PROGRESS
        existing.save()
        messages.info(request, f"You have been re-enrolled in {course.name}.")
        return redirect("home")

    # Check enrollment limit (max 5 courses per cohort)
    current_enrollment_count = Enrollment.objects.filter(
        registration=registration
    ).count()
    if current_enrollment_count >= 5:
        messages.error(
            request,
            f"You cannot enroll in more than 5 courses in the {cohort.name} cohort.",
        )
        return redirect("home")

    # Create enrollment
    Enrollment.objects.create(
        course=course, registration=registration, status="IN_PROGRESS"
    )

    messages.success(request, f"Successfully enrolled in {course.name}!")
    return redirect("home")


# @login_required
# def unenroll_from_course(request, enrollment_id):
#     """Allow student to unenroll from a course"""
#     if request.method != 'POST':
#         return redirect('home')
#
#     # Get student
#     try:
#         student = Student.objects.get(user=request.user)
#     except Student.DoesNotExist:
#         messages.error(request, 'Student profile not found.')
#         return redirect('home')
#
#     # Get enrollment and verify it belongs to this student
#     enrollment = get_object_or_404(Enrollment, id=enrollment_id, student=student)
#
#     course_name = enrollment.course.name
#
#     # Delete the enrollment
#     enrollment.delete()
#
#     messages.success(request, f'Successfully unenrolled from {course_name}.')
#     return redirect('home')
# Disabled: Students cannot unenroll from courses


@login_required
def my_certificates(request):
    """Display certificates for the logged-in student"""
    # Get student
    try:
        student = Student.objects.get(user=request.user)
    except Student.DoesNotExist:
        messages.error(request, "Student profile not found.")
        return redirect("home")

    # Get all enrollments with certificate info
    enrollments = (
        Enrollment.objects.filter(student=student)
        .select_related("course", "cohort", "registration", "certificate")
        .order_by("-cohort__start_date", "course__name")
    )

    # Build data structure with certificates and eligibility info
    certificates_data = []
    for enrollment in enrollments:
        # Get certificate if it exists (using the OneToOne relationship)
        course_cert = (
            enrollment.certificate if hasattr(enrollment, "certificate") else None
        )

        certificates_data.append(
            {
                "enrollment": enrollment,
                "course": enrollment.course,
                "cohort": enrollment.cohort,
                "is_eligible": enrollment.certificate_eligible,
                "eligibility_notes": enrollment.certificate_eligibility_notes,
                "certificate": course_cert,
                "completion_rate": enrollment.completion_rate,
                "average_score": enrollment.overall_average_score,
            }
        )

    context = {
        "student": student,
        "certificates_data": certificates_data,
    }
    return render(request, "app/my_certificates.html", context)


@staff_member_required
def issue_certificate(request, enrollment_id):
    """Issue a certificate for an enrollment"""
    from datetime import date

    if request.method != "POST":
        messages.error(request, "Invalid request method.")
        return redirect("home")

    # Get the enrollment
    enrollment = get_object_or_404(
        Enrollment.objects.select_related(
            "registration__student", "course", "registration"
        ),
        id=enrollment_id,
    )

    # Check eligibility
    if not enrollment.certificate_eligible:
        messages.error(
            request,
            f"Student is not eligible for certificate for {enrollment.course.name}. {enrollment.certificate_eligibility_notes}",
        )
        return redirect(request.META.get("HTTP_REFERER", "home"))

    # Check if certificate already exists
    if hasattr(enrollment, "certificate"):
        messages.warning(
            request,
            f"Certificate already issued for {enrollment.student.full_name} - {enrollment.course.name}",
        )
        return redirect(request.META.get("HTTP_REFERER", "home"))

    # Create certificate
    try:
        from ..services import generate_certificate

        cert = Certificate.objects.create(
            enrollment=enrollment,
            issued_date=date.today(),
            completion_percentage=enrollment.completion_rate or 0,
            average_grade=enrollment.overall_average_score,
            notes=f"Issued by {request.user.username}",
        )

        # Generate and save certificate file
        certificate_file = generate_certificate(cert)
        cert.certificate_file.save(certificate_file.name, certificate_file, save=True)

        messages.success(
            request,
            f"Certificate issued for {enrollment.student.full_name} - {enrollment.course.name}",
        )
    except Exception as e:
        messages.error(request, f"Error issuing certificate: {str(e)}")

    return redirect(request.META.get("HTTP_REFERER", "home"))


@staff_member_required
def delete_certificate(request, certificate_id):
    """Delete a certificate record (staff only)"""
    if request.method != "POST":
        messages.error(request, "Invalid request method.")
        return redirect("home")

    # Get the certificate
    certificate = get_object_or_404(Certificate, id=certificate_id)

    student_name = certificate.registration.student.full_name
    course_name = certificate.course.name if certificate.course else "Cohort"

    try:
        # Delete the file if it exists
        if certificate.certificate_file:
            certificate.certificate_file.delete(save=False)

        # Delete the certificate record
        certificate.delete()

        messages.success(
            request, f"Certificate deleted for {student_name} - {course_name}"
        )
    except Exception as e:
        messages.error(request, f"Error deleting certificate: {str(e)}")

    return redirect(request.META.get("HTTP_REFERER", "home"))


def test_certificate(request):
    """Test view to preview certificate template"""
    context = {
        "name": "Jane Doe",
        "course": "Digital Marketing Fundamentals",
        "period": "January 2026 - May 2026",
    }

    return render(request, "certificate/certificate.html", context)


def view_certificate(request, student_google_id, course_google_id):
    """View a student's certificate for a specific course"""
    from django.shortcuts import get_object_or_404

    # Get the student and course
    student = get_object_or_404(Student, google_id=student_google_id)
    course = get_object_or_404(Course, google_id=course_google_id)

    # Find the certificate for this student and course
    certificate = (
        Certificate.objects.filter(
            enrollment__registration__student=student, enrollment__course=course
        )
        .select_related(
            "enrollment__registration__student",
            "enrollment__course",
            "enrollment__registration__cohort",
        )
        .first()
    )

    if not certificate:
        messages.error(request, "Certificate not found")
        return redirect("home")

    # Get cohort dates
    cohort = certificate.enrollment.cohort
    period = (
        f"{cohort.start_date.strftime('%B %Y')} - {cohort.end_date.strftime('%B %Y')}"
    )

    context = {
        "name": student.full_name,
        "course": certificate.enrollment.course.name,
        "period": period,
    }

    return render(request, "certificate/certificate.html", context)


# =============================================================================
# STAFF VIEWS FOR REGISTRATIONS
# =============================================================================
